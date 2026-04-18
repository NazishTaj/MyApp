from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .utils import has_permission
from datetime import date , datetime
from django.core.paginator import Paginator
from .forms import ClinicScheduleForm
import openpyxl
from django.http import HttpResponse
from .models import Patient, Appointment, Prescription, UserProfile , ClinicSchedule , Clinic
from django.core.exceptions import ValidationError
from .models import Bill, BillItem, UserProfile
from django.db.models import Sum
from django.utils import timezone

def clinic_blocked(request):
    return render(request, 'clinic_blocked.html')

def user_blocked(request):
    return render(request, 'user_blocked.html')


def format_medicines(request):
    names = request.POST.getlist('medicine_name[]')
    dosages = request.POST.getlist('dosage[]')
    durations = request.POST.getlist('duration[]')
    remarks = request.POST.getlist('remark[]')

    med_list = []

    for i in range(len(names)):
        name = (names[i] or "").strip()
        dose = (dosages[i] if i < len(dosages) else "").strip()
        duration = (durations[i] if i < len(durations) else "").strip()
        remark = (remarks[i] if i < len(remarks) else "").strip()

        # 👇 allow partial entries (safe)
        if name or dose or duration or remark:
            duration = (durations[i] if i < len(durations) else "").strip()
            med_list.append(f"{name}||{dose}||{duration}||{remark}")

    return "\n".join(med_list)


def billing_enabled(request):

    # ✅ check login
    if not request.user.is_authenticated:
        return False

    profile = get_object_or_404(UserProfile, user=request.user)

    return profile.clinic.billing_enabled


# ---------------- LOGIN ---------------- #

def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            return redirect("dashboard")

        return render(request, "login.html", {"error": "Invalid username or password"})

    return render(request, "login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


# ---------------- DASHBOARD ---------------- #

from django.utils import timezone
@login_required(login_url="login")
def dashboard(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    doctors = UserProfile.objects.filter(
        clinic=clinic,
        role__in=["owner", "doctor"]
    )

    show_doctor_column = (
        doctors.count() > 1 and profile.role == "receptionist"
    )

    today = timezone.localtime().date()
    if profile.role in ["doctor", "owner"]:
        appointments_today = Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            doctor=profile
        )

    elif profile.role == "assistant":
        if profile.assigned_doctor:
            appointments_today = Appointment.objects.filter(
                clinic=clinic,
                appointment_date=today,
                doctor=profile.assigned_doctor
            )
        else:
            appointments_today = Appointment.objects.none()

    else:  # receptionist
        appointments_today = Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today
        )
    appointments_today = appointments_today.order_by(
        "doctor", "token_number"
    )


    today_revenue = Bill.objects.filter(
        clinic=clinic,
        created_at__date=today
    ).aggregate(total=Sum("total_amount"))["total"] or 0

    today_appointments = appointments_today.count()
    pending_appointments = appointments_today.filter(status="pending").count()
    completed_appointments = appointments_today.filter(status="completed").count()

    total_patients = Patient.objects.filter(clinic=clinic,is_active=True).count()
    total_appointments = Appointment.objects.filter(clinic=clinic).count()
    from collections import defaultdict  # (top pe import kar lena ek baar)

    # 🔥 busy doctors find karo
    busy_doctors = set(
        Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            queue_status="in_consultation"
        ).values_list("doctor_id", flat=True)
    )

    next_tokens = {}

    all_waiting = Appointment.objects.filter(
        clinic=clinic,
        appointment_date=today,
        queue_status="waiting"
    ).order_by("doctor", "token_number")

    seen_doctors = set()

    for appt in all_waiting:
        if appt.doctor_id not in seen_doctors and appt.doctor_id not in busy_doctors:
            next_tokens[appt.id] = True
            seen_doctors.add(appt.doctor_id)

    context = {
        "appointments": appointments_today,
        "show_doctor_column": show_doctor_column,
        "total_patients": total_patients,
        "total_appointments": total_appointments,
        "pending_appointments": pending_appointments,
        "completed_appointments": completed_appointments,
        "today_appointments": today_appointments,
        "today_revenue": today_revenue,
        "next_tokens": next_tokens
    }

    response = render(request, "dashboard.html", context)

    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"

    return response
# ---------------- PATIENTS ---------------- #

from django.db.models import Q
from django.contrib.auth.decorators import login_required

@login_required(login_url="login")
def patient_list(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    query = request.GET.get("q", "").strip()

    patients_list = Patient.objects.filter(clinic=clinic,is_active=True).order_by('patient_id')

    if query:
        patients_list = patients_list.filter(
            Q(name__icontains=query) |
            Q(phone__icontains=query)
        )

    # 🔥 pagination logic
    per_page = request.GET.get("limit", 20)

    try:
        per_page = int(per_page)
    except:
        per_page = 20

    if per_page not in [20, 50]:
        per_page = 20

    paginator = Paginator(patients_list, per_page)

    page = request.GET.get("page")

    patients = paginator.get_page(page)

    context = {
        "patients": patients,
        "query": query,
        "limit": per_page
    }

    return render(request, "patients.html", context)


from django.core.exceptions import ValidationError

@login_required(login_url="login")
def add_patient(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not has_permission(request.user, "manage_patients"):
        return render(request, "403.html", status=403)

    prefill_phone = request.GET.get("phone","")

    if request.method == "POST":

        name = request.POST["name"]
        phone = request.POST["phone"]
        age = request.POST["age"]
        gender = request.POST["gender"]
        address = request.POST.get("address")

        force_create = request.POST.get("force_create")

        existing_patient = Patient.objects.filter(
            clinic=clinic,
            phone=phone
        ).first()

        # 🔴 Duplicate detected but user didn't confirm
        if existing_patient and not force_create:
            return render(
                request,
                "add_patient.html",
                {
                    "error_duplicate": "Patient with this phone number already exists.",
                    "prefill_phone": phone,
                    "name": name,
                    "age": age,
                    "gender": gender,
                    "address": address,
                    "show_continue": True
                }
            )

        patient = Patient(
            clinic=clinic,
            name=name,
            phone=phone,
            age=age,
            gender=gender,
            address=address
        )

        try:
            patient.full_clean()
            patient.save()
            return redirect("patient_list")

        except ValidationError as e:

            return render(
                request,
                "add_patient.html",
                {
                    "error": e.message_dict,
                    "prefill_phone": phone
                }
            )

    return render(
        request,
        "add_patient.html",
        {"prefill_phone": prefill_phone}
    )
@login_required(login_url="login")
def edit_patient(request, patient_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not has_permission(request.user, "manage_patients"):
        return render(request, "403.html", status=403)

    patient = get_object_or_404(Patient, id=patient_id, clinic=clinic,is_active=True)

    if request.method == "POST":

        patient.name = request.POST.get("name")
        patient.phone = request.POST.get("phone")
        patient.age = request.POST.get("age")
        patient.gender = request.POST.get("gender")
        patient.address = request.POST.get("address")

        patient.save()

        return redirect("patient_list")

    return render(request, "edit_patient.html", {"patient": patient})


@login_required(login_url="login")
def delete_patient(request, patient_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not has_permission(request.user, "manage_patients"):
        return render(request, "403.html", status=403)

    patient = get_object_or_404(Patient, id=patient_id, clinic=clinic)

    patient.is_active = False
    patient.save()

    return redirect("patient_list")


# ---------------- APPOINTMENTS ---------------- #


from django.utils import timezone


@login_required(login_url="login")
def book_appointment(request, patient_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not has_permission(request.user, "manage_appointments"):
        return render(request, "403.html", status=403)

    patient = get_object_or_404(Patient, id=patient_id, clinic=clinic,is_active=True)

    if request.method == "POST":

        date_val = request.POST.get("date")
        time = request.POST.get("time")
        problem = request.POST.get("problem")
        visit_type = request.POST.get("visit_type") or "new"
        payment_mode = request.POST.get("payment_mode")
        doctors = UserProfile.objects.filter(
            clinic=clinic,
        role__in=["owner", "doctor"]
        )

        # 🔥 ROLE BASED DOCTOR LOGIC
        if profile.role in ["doctor", "owner"]:
            doctor = profile   # auto assign logged-in doctor

        else:
            doctor_id = request.POST.get("doctor_id")

            if not doctor_id:
                return render(request, "book_appointment.html", {
                    "patient": patient,
                    "doctors": doctors,
                    "error": "Please select doctor",
                    "billing_enabled": clinic.billing_enabled,
                    "profile": profile   # 👈 ADD
                })

            doctor = UserProfile.objects.filter(
                id=doctor_id,
                clinic=clinic,
                role__in=["owner", "doctor"]
            ).first()

            if not doctor:
                return render(request, "book_appointment.html", {
                    "patient": patient,
                    "doctors": doctors,
                    "error": "Invalid doctor selected",
                    "billing_enabled": clinic.billing_enabled,
                    "profile": profile   # 👈 ADD
                }) 
        posted_fee = request.POST.get("consultation_fee")
        fee = float(posted_fee) if posted_fee not in [None, ""] else (doctor.consultation_fee or 0)

        # ✅ DATE FIX (MAIN FIX)
        if date_val:
            date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
        else:
            date_val = timezone.localtime().date()

        # ✅ TIME FIX
        if time:
            time = datetime.strptime(time, "%H:%M").time()
        else:
            time = timezone.localtime().time()


   
        # ✅ Payment logic (FIXED)
        if visit_type == "free":
            fee = 0
            payment_status = "waived"
            payment_mode = "free"   # ✅ FIX

        else:
            if clinic.billing_enabled:
                payment_status = "paid" if fee > 0 else "unpaid"
            else:
                payment_status = "unpaid"

            # ✅ IMPORTANT FIX
            if fee == 0:
                payment_mode = "free"
            else:
                payment_mode = request.POST.get("payment_mode") or "cash"

  

        # 🔹 Token logic
        last_token = Appointment.objects.filter(
            clinic=clinic,
            appointment_date=date_val,
            doctor=doctor
        ).order_by('-token_number').first()

        if last_token and last_token.token_number:
            token = last_token.token_number + 1
        else:
            token = 1

        appointment = Appointment.objects.create(
            clinic=clinic,
            patient=patient,
            appointment_date=date_val,
            appointment_time=time,
            problem=problem,
            token_number=token,
            doctor=doctor,
            visit_type=visit_type,
            consultation_fee=fee, 
            payment_status=payment_status,
            payment_mode=payment_mode,
            status="pending",
            queue_status="waiting"
        )
        if clinic.billing_enabled:

            bill = Bill.objects.create(
                clinic=clinic,
                patient=patient,
                doctor=doctor,
                appointment=appointment,
                total_amount=appointment.consultation_fee, 
                payment_mode=payment_mode
            )

            BillItem.objects.create(
                bill=bill,
                item_name="Consultation",
                amount=appointment.consultation_fee
            )
        messages.success(
        request,
        f"{patient.name} checked-in successfully (Token #{token}) ✅"
        )

        return redirect("patient_list")   # 🔥 IMPORTANT

    doctors = UserProfile.objects.filter(
        clinic=clinic,
        role__in=["owner", "doctor"]
    )

    return render(request, "book_appointment.html", {
        "patient": patient,
        "doctors": doctors,
        "billing_enabled": clinic.billing_enabled,
        "profile": profile
    })




@login_required(login_url="login")
def appointments(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not has_permission(request.user, "manage_appointments"):
        return render(request, "403.html", status=403)

    # 🔥 NEW LOGIC
    if profile.role == "doctor":
        appointments = Appointment.objects.filter(
            clinic=clinic,
            doctor=profile
        )

    elif profile.role == "assistant":
        if profile.assigned_doctor:
            appointments = Appointment.objects.filter(
                clinic=clinic,
                doctor=profile.assigned_doctor
            )
        else:
            appointments = Appointment.objects.none()

    else:
        appointments = Appointment.objects.filter(clinic=clinic)

    

    # ========== FILTERS ==========
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    patient_id = request.GET.get('patient')
    doctor_id = request.GET.get('doctor')
    status = request.GET.get('status')
    
    if start_date:
        appointments = appointments.filter(appointment_date__gte=start_date)
    
    if end_date:
        appointments = appointments.filter(appointment_date__lte=end_date)
    
    if patient_id:
        appointments = appointments.filter(patient_id=patient_id)
    
    if doctor_id:
        appointments = appointments.filter(doctor_id=doctor_id)
    
    if status:
        appointments = appointments.filter(status=status)

    appointments = appointments.order_by("-appointment_date", "appointment_time")
    # ✅ pagination
    per_page = request.GET.get("limit", 20)

    try:
        per_page = int(per_page)
    except:
        per_page = 20

    if per_page not in [20, 50]:
        per_page = 20

    paginator = Paginator(appointments, per_page)
    page = request.GET.get("page")
    appointments = paginator.get_page(page)

    # ========== DROPDOWN DATA ==========
    patients = Patient.objects.filter(clinic=clinic).order_by('name')
    
    doctors = UserProfile.objects.filter(
        clinic=clinic,
        role__in=["owner", "doctor"]
    ).order_by('name')
    show_doctor_column = doctors.count() > 1

    return render(request, "appointments.html", {
        "appointments": appointments,
        "show_doctor_column": show_doctor_column,
        "patients": patients,
        "doctors": doctors,
         "limit": per_page
    })


@login_required(login_url="login")
def complete_appointment(request, appointment_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not has_permission(request.user, "manage_appointments"):
        return render(request, "403.html", status=403)

    appointment = get_object_or_404(Appointment, id=appointment_id, clinic=clinic)

    appointment.status = "completed"
    appointment.queue_status = "done"
    appointment.save()

    # 🔥 NEXT TOKENS
    today = appointment.appointment_date

    busy_doctors = set(
        Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            queue_status="in_consultation"
        ).values_list("doctor_id", flat=True)
    )

    all_waiting = Appointment.objects.filter(
        clinic=clinic,
        appointment_date=today,
        queue_status="waiting"
    ).order_by("doctor", "token_number")

    next_tokens = []
    seen_doctors = set()

    for appt in all_waiting:
        if appt.doctor_id not in seen_doctors and appt.doctor_id not in busy_doctors:
            next_tokens.append(appt.id)
            seen_doctors.add(appt.doctor_id)

    # 🔥 WEBSOCKET
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync

    channel_layer = get_channel_layer()

    group_name = f"dashboard_{clinic.id}"

    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            "type": "send_update",
            "data": {
                "appointment_id": appointment.id,
                "patient_id": appointment.patient.id,
                "status": appointment.status,
                "queue_status": appointment.queue_status,
                "next_tokens": next_tokens,
            }
        }
    )

    from django.http import JsonResponse
    return JsonResponse({"status": "ok"})


@login_required(login_url="login")
def send_to_doctor(request, appointment_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    appointment = get_object_or_404(Appointment, id=appointment_id, clinic=clinic)

    if profile.role != "receptionist" or not clinic.is_advanced:
        return redirect("dashboard")

    Appointment.objects.filter(
        clinic=clinic,
        doctor=appointment.doctor,
        queue_status="in_consultation"
    ).update(queue_status="waiting")

    appointment.queue_status = "in_consultation"
    appointment.save()

    # 🔥 NEXT TOKENS
    today = appointment.appointment_date

    busy_doctors = set(
        Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            queue_status="in_consultation"
        ).values_list("doctor_id", flat=True)
    )

    all_waiting = Appointment.objects.filter(
        clinic=clinic,
        appointment_date=today,
        queue_status="waiting"
    ).order_by("doctor", "token_number")

    next_tokens = []
    seen_doctors = set()

    for appt in all_waiting:
        if appt.doctor_id not in seen_doctors and appt.doctor_id not in busy_doctors:
            next_tokens.append(appt.id)
            seen_doctors.add(appt.doctor_id)

    # 🔥 WEBSOCKET
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync

    channel_layer = get_channel_layer()
    
    group_name = f"dashboard_{clinic.id}"
    async_to_sync(channel_layer.group_send)(
        group_name,

        {
            "type": "send_update",
            "data": {
                "appointment_id": appointment.id,
                "patient_id": appointment.patient.id,
                "status": appointment.status,
                "queue_status": appointment.queue_status,
                "next_tokens": next_tokens,
            }
        }
    )

    from django.http import JsonResponse
    return JsonResponse({"status": "ok"})


@login_required(login_url="login")
def cancel_appointment(request, appointment_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not has_permission(request.user, "manage_appointments"):
        return render(request, "403.html", status=403)

    appointment = get_object_or_404(Appointment, id=appointment_id, clinic=clinic)

    appointment.status = "cancelled"
    appointment.queue_status = "done"
    appointment.save()

    # 🔥 NEXT TOKENS
    today = appointment.appointment_date

    busy_doctors = set(
        Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            queue_status="in_consultation"
        ).values_list("doctor_id", flat=True)
    )

    all_waiting = Appointment.objects.filter(
        clinic=clinic,
        appointment_date=today,
        queue_status="waiting"
    ).order_by("doctor", "token_number")

    next_tokens = []
    seen_doctors = set()

    for appt in all_waiting:
        if appt.doctor_id not in seen_doctors and appt.doctor_id not in busy_doctors:
            next_tokens.append(appt.id)
            seen_doctors.add(appt.doctor_id)

    # 🔥 WEBSOCKET
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync

    channel_layer = get_channel_layer()

    group_name = f"dashboard_{clinic.id}"

    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            "type": "send_update",
            "data": {
                "appointment_id": appointment.id,
                "patient_id": appointment.patient.id,
                "status": appointment.status,
                "queue_status": appointment.queue_status,
                "next_tokens": next_tokens,
            }
        }
    )

    from django.http import JsonResponse
    return JsonResponse({"status": "ok"})

# ---------------- PRESCRIPTIONS ---------------- #

@login_required(login_url="login")
def add_prescription(request, patient_id):

    if not has_permission(request.user, "create_prescription"):
        return render(request, "403.html", status=403)

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    patient = get_object_or_404(Patient, id=patient_id, clinic=clinic,is_active=True)

    if request.method == "POST":

        diagnosis = request.POST.get("diagnosis")
        symptoms = request.POST.get("symptoms")
        medicines = format_medicines(request)
        # 🔥 AUTO SAVE MEDICINES (yahi daalna hai)
        med_lines = medicines.split("\n")

        for med in med_lines:
            med = med.strip()

            if not med:
                continue

            # name extract
            if "||" in med:
                name = med.split("||")[0]
            else:
                name = med

            save_medicine(name, clinic)
        tests = request.POST.get("tests")
        notes = request.POST.get("notes")
        weight = request.POST.get("weight")
        blood_group = request.POST.get("blood_group")

        # 🔥 NEW DOCTOR LOGIC (FINAL)
        doctor_id = request.POST.get("doctor_id")

        if profile.role == "owner":
            doctor = profile

        elif profile.role == "doctor":
            doctor = profile

        elif profile.role == "assistant":
            doctor = profile.assigned_doctor

        elif profile.role == "receptionist":
            if doctor_id:
                doctor = UserProfile.objects.filter(
                    id=doctor_id,
                    clinic=clinic,
                    role__in=["owner", "doctor"]
                ).first()
            else:
                doctor = None
        else:
            doctor = None

        # 🔥 CREATE PRESCRIPTION
        Prescription.objects.create(
            clinic=clinic,
            patient=patient,
            diagnosis=diagnosis,
            symptoms=symptoms,
            medicines=medicines,
            tests=tests,
            notes=notes,
            weight=weight,
            blood_group=blood_group,
            created_by=profile,
            doctor=doctor
        )

        return redirect("patient_history", patient_id=patient.id)

    # 🔥 Doctors list (for receptionist dropdown)
    doctors = UserProfile.objects.filter(
        clinic=clinic,
        role__in=["owner", "doctor"]
    )

    return render(request, "add_prescription.html", {
        "patient": patient,
        "doctors": doctors,
        "profile": profile 
    })
    
#Revise Prescription

@login_required(login_url="login")
def revise_prescription(request, id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    old = get_object_or_404(
        Prescription.objects.select_related("patient", "doctor"),
        id=id,
        clinic=clinic
    )

    patient = old.patient

    if request.method == "POST":

        diagnosis = request.POST.get("diagnosis")
        symptoms = request.POST.get("symptoms")
        medicines = format_medicines(request)
        tests = request.POST.get("tests")
        notes = request.POST.get("notes")
        weight = request.POST.get("weight")
        blood_group = request.POST.get("blood_group")

        doctor = old.doctor

        Prescription.objects.create(
            clinic=clinic,
            patient=patient,
            diagnosis=diagnosis,
            symptoms=symptoms,
            medicines=medicines,
            tests=tests,
            notes=notes,
            weight=weight,
            blood_group=blood_group,
            created_by=profile,
            doctor=doctor,
            parent=old
        )

        return redirect("patient_history", patient_id=patient.id)

    # 🔥 NEW: medicines prefill logic
    med_lines = []

    if old.medicines:
        for med in old.medicines.split("\n"):
            if "||" in med:
                parts = med.split("||")
                med_lines.append({
                    "name": parts[0],
                    "dose": parts[1] if len(parts) > 1 else "",
                    "duration": parts[2] if len(parts) > 2 else "",
                    "remark": parts[3] if len(parts) > 3 else "",
                })
            else:
                med_lines.append({
                    "name": med,
                    "dose": "",
                    "duration": "",
                    "remark": "",
                })

    return render(request, "add_prescription.html", {
        "patient": patient,
        "profile": profile,
        "doctors": UserProfile.objects.filter(
            clinic=clinic,
            role__in=["owner", "doctor"]
        ),
        "old": old,
        "med_lines": med_lines   # 🔥 IMPORTANT FIX
    })


@login_required(login_url="login")
def patient_history(request, patient_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
   
    patient = get_object_or_404(Patient, id=patient_id, clinic=clinic)

    bills = Bill.objects.filter(
        clinic=clinic,
        patient=patient
    ).order_by("-created_at")

    prescriptions = Prescription.objects.filter(
        clinic=clinic,
        patient=patient
    ).select_related("doctor", "patient").prefetch_related("revisions").order_by("-created_at")

    # ✅ FIXED: medicines attach to each prescription
    for p in prescriptions:
        med_lines = []

        if p.medicines:
            for med in p.medicines.split("\n"):
                if "||" in med:
                    parts = med.split("||")
                    med_lines.append({
                        "name": parts[0],
                        "dose": parts[1] if len(parts) > 1 else "",
                        "duration": parts[2] if len(parts) > 2 else "",
                        "remark": parts[3] if len(parts) > 3 else "",
                    })
                else:
                    med_lines.append({
                        "name": med,
                        "dose": "",
                        "duration": "",
                        "remark": "",
                    })

        p.med_lines = med_lines  # 🔥 attach here

    return render(request, "patient_history.html", {
        "patient": patient,
        "prescriptions": prescriptions,
        "bills": bills
    })


@login_required
def create_bill_for_patient(request, patient_id):
    if not has_permission(request.user, "manage_billing"):
        return render(request, "403.html", status=403)
    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not clinic.billing_enabled:
        return redirect("dashboard")

    patient = get_object_or_404(Patient, id=patient_id, clinic=clinic,is_active=True)

    if request.method == "POST":

        payment_mode = request.POST.get("payment_mode")
        discount_percent = float(request.POST.get("discount")or 0)

        doctor = None
        # Create bill
        bill = Bill.objects.create(
            clinic=clinic,
            patient=patient,   # 🔥 FIXED PATIENT
            doctor=doctor,
            payment_mode=payment_mode,
        )

        item_names = request.POST.getlist("item_name[]")
        item_amounts = request.POST.getlist("item_amount[]")

        subtotal = 0

        for name, amount in zip(item_names, item_amounts):
            if name and amount:
                amount = float(amount)

                BillItem.objects.create(
                    bill=bill,
                    item_name=name,
                    amount=amount
                )

                subtotal += amount

        discount_amount = (subtotal * discount_percent) / 100
        total = subtotal - discount_amount

        bill.subtotal = subtotal
        bill.discount = discount_percent
        bill.discount_amount = discount_amount
        bill.total_amount = total
        bill.save()

        return redirect("view_bill", bill_id=bill.id)

    return render(request, "billing/create_bill.html", {
        "patient": patient,
        "clinic": clinic,
        "hide_watermark": True
    })

@login_required(login_url="login")
def view_prescription(request, id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not has_permission(request.user, "create_prescription"):
        return render(request, "403.html", status=403)

    prescription = get_object_or_404(
        Prescription.objects.select_related("patient", "doctor", "clinic"),
        id=id,
        clinic=clinic
    )

    # 👇 NEW CODE
    med_lines = []

    if prescription.medicines:
        for med in prescription.medicines.split("\n"):
            if "||" in med:
                parts = med.split("||")
                med_lines.append({
                    "name": parts[0],
                    "dose": parts[1] if len(parts) > 1 else "",
                    "duration": parts[2] if len(parts) > 2 else "",
                    "remark": parts[3] if len(parts) > 3 else "",
                })
            else:
                med_lines.append({
                    "name": med,
                    "dose": "",
                    "remark": "",
                })

    return render(request, "view_prescription.html", {
        "prescription": prescription,
        "clinic": clinic,
        "med_lines": med_lines,
        "hide_watermark": True
    })


# ---------------- ONLINE BOOKING ---------------- #

def online_booking(request):

    if request.method == "POST":

        name = request.POST["name"]
        phone = request.POST["phone"]
        age = request.POST["age"]
        date_val = request.POST["date"]
        time = request.POST["time"]
        problem = request.POST["problem"]

        clinic = Clinic.objects.first()

        if not clinic:
            return HttpResponse("No clinic configured", status=400)
            

        # ✅ DATE FIX
        if date_val:
            date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
        else:
            date_val = timezone.localtime().date()

        # ✅ TIME FIX
        if time:
            time = datetime.strptime(time, "%H:%M").time()
        else:
            time = datetime.now().time()

        # ✅ PATIENT (avoid duplicate)
        patient, created = Patient.objects.get_or_create(
            clinic=clinic,
            phone=phone,
            defaults={
                "name": name,
                "age": age,
                "gender": "Unknown"
            }
        )
        doctor = UserProfile.objects.filter(
            clinic=clinic,
            role__in=["owner", "doctor"]
        ).first()

        # ✅ PREVENT DUPLICATE BOOKING SAME DAY
        if Appointment.objects.filter(
            clinic=clinic,
            patient=patient,
            appointment_date=date_val
        ).exists():
            return render(request, "booking_success.html", {
                "message": "You already booked for this date"
            })

        # ✅ TOKEN LOGIC (per day reset)
        last_token = Appointment.objects.filter(
            clinic=clinic,
            appointment_date=date_val,
            doctor=doctor
        ).order_by('-token_number').first()

        token = last_token.token_number + 1 if last_token else 1

        # ✅ CREATE APPOINTMENT
        Appointment.objects.create(
            clinic=clinic,
            patient=patient,
            appointment_date=date_val,
            appointment_time=time,
            problem=problem,
            token_number=token,
            doctor=doctor,
            queue_status="waiting"
        )

        return render(request, "booking_success.html", {
            "message": "Booking successful"
        })

    return render(request, "online_booking.html")




@login_required(login_url="login")
def profile(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    context = {
        "user": request.user,
        "profile": profile,
        "clinic": clinic
    }

    return render(request, "profile.html", context)



#edit profile

@login_required
def edit_profile(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    schedules = ClinicSchedule.objects.filter(clinic=clinic).order_by('day','start_time')

    if request.method == "POST":

        profile.name = request.POST.get("name")
        profile.phone = request.POST.get("phone")

        # ✅ FIX ADDED
        if profile.role in ["doctor", "owner"]:
            profile.degree = request.POST.get("degree") or ""
            profile.reg_no = request.POST.get("reg_no") or ""
            profile.consultation_fee = float(request.POST.get("consultation_fee") or 0)

        if profile.is_owner:
            clinic.name = request.POST.get("clinic_name")
            clinic.phone = request.POST.get("clinic_phone")
            clinic.address = request.POST.get("clinic_address")

            clinic.billing_enabled = bool(request.POST.get("billing_enabled"))
            clinic.is_advanced = bool(request.POST.get("is_advanced"))

        if request.FILES.get("photo"):
            profile.photo = request.FILES["photo"]

        if request.FILES.get("logo"):
            clinic.logo = request.FILES["logo"]

        profile.save()
        clinic.save()

        return redirect("profile")

    return render(request, "edit_profile.html", {
        "profile": profile,
        "clinic": clinic,
        "schedules": schedules
    })


@login_required
def add_schedule(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if request.method == "POST":
        form = ClinicScheduleForm(request.POST)

        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.clinic = clinic
            schedule.save()

            print("Schedule saved:", schedule.day, schedule.start_time, schedule.end_time)

            return redirect("profile")

        else:
            print("FORM ERRORS:", form.errors)  # debug ke liye

    else:
        form = ClinicScheduleForm()

    return render(request, "add_schedule.html", {
        "form": form
    })

@login_required
def delete_schedule(request, id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    schedule = ClinicSchedule.objects.get(id=id, clinic=clinic)
    schedule.delete()

    return redirect("edit_profile")

@login_required
def edit_schedule(request, id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    schedule = ClinicSchedule.objects.get(id=id, clinic=clinic)

    if request.method == "POST":

        schedule.day = request.POST.get("day")
        schedule.start_time = request.POST.get("start_time")
        schedule.end_time = request.POST.get("end_time")

        schedule.save()

        return redirect("edit_profile")

    return render(request, "edit_schedule.html", {
        "schedule": schedule
    })

#Excel Export

@login_required
def export_month_appointments(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not profile.is_owner:
        return render(request, "403.html", status=403)

    month = request.GET.get("month")
    year = request.GET.get("year")

    if not month or not year:
        today = date.today()
        month = today.month
        year = today.year

    appointments = Appointment.objects.filter(
        clinic=clinic,
        appointment_date__month=int(month),
        appointment_date__year=int(year)
    ).select_related("patient")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Appointments"

    headers = [
        "Patient Name",
        "Phone",
        "Age",
        "Gender",
        "Address",
        "Date",
        "Time",
        "Problem",
        "Status"
    ]

    sheet.append(headers)

    for appt in appointments:

        patient = appt.patient

        sheet.append([
            patient.name,
            patient.phone,
            patient.age,
            patient.gender,
            patient.address,
            appt.appointment_date.strftime("%d-%m-%Y"),
            appt.appointment_time.strftime("%H:%M"),
            appt.problem,
            appt.status
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"appointments_{month}_{year}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response



@login_required
def export_all_appointments(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not profile.is_owner:
        return render(request, "403.html", status=403)

    appointments = Appointment.objects.filter(clinic=clinic)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Appointments"

    headers = [
        "Patient Name",
        "Phone",
        "Age",
        "Gender",
        "Address",
        "Date",
        "Time",
        "Problem",
        "Status"
    ]

    sheet.append(headers)

    for appt in appointments:

        patient = appt.patient   # <-- important line

        sheet.append([
            patient.name,
            patient.phone,
            patient.age,
            patient.gender,
            patient.address,
            appt.appointment_date.strftime("%d-%m-%Y"),
            appt.appointment_time.strftime("%H:%M"),
            appt.problem,
            appt.status
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename="all_appointments.xlsx"'

    workbook.save(response)

    return response


#------------------------------------------------------------


@login_required
def mark_pending(request, appointment_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not has_permission(request.user, "manage_appointments"):
        return render(request, "403.html", status=403)

    appointment = get_object_or_404(Appointment, id=appointment_id, clinic=clinic)

    appointment.status = "pending"
    appointment.queue_status = "waiting"
    appointment.save()

    # 🔥 NEXT TOKENS
    today = appointment.appointment_date

    busy_doctors = set(
        Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            queue_status="in_consultation"
        ).values_list("doctor_id", flat=True)
    )

    all_waiting = Appointment.objects.filter(
        clinic=clinic,
        appointment_date=today,
        queue_status="waiting"
    ).order_by("doctor", "token_number")

    next_tokens = []
    seen_doctors = set()

    for appt in all_waiting:
        if appt.doctor_id not in seen_doctors and appt.doctor_id not in busy_doctors:
            next_tokens.append(appt.id)
            seen_doctors.add(appt.doctor_id)

    # 🔥 WEBSOCKET
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync

    channel_layer = get_channel_layer()

    group_name = f"dashboard_{clinic.id}"

    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            "type": "send_update",
            "data": {
                "appointment_id": appointment.id,
                "patient_id": appointment.patient.id,
                "status": appointment.status,
                "queue_status": appointment.queue_status,
                "next_tokens": next_tokens,
            }
        }
    )

    from django.http import JsonResponse
    return JsonResponse({"status": "ok"})



#Bill 
@login_required
def create_bill(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not has_permission(request.user, "manage_billing"):
        return render(request, "403.html", status=403)

    if not clinic.billing_enabled:
        return redirect("dashboard")

    patients = Patient.objects.filter(clinic=clinic,is_active=True)
    doctors = UserProfile.objects.filter(
        clinic=clinic,
        role__in=["owner", "doctor"]
    )

    if request.method == "POST":

        patient_id = request.POST.get("patient")
        referred_by_id = request.POST.get("referred_by")
        payment_mode = request.POST.get("payment_mode", "").strip().lower()
        discount_percent = float(request.POST.get("discount") or 0)

        patient = get_object_or_404(Patient, id=patient_id, clinic=clinic,is_active=True)
   

        # 🔥 Bill number generate
        last_bill = Bill.objects.filter(clinic=clinic).order_by("-id").first()

        if last_bill:
            last_number = int(last_bill.bill_number.split("-")[1])
            new_number = last_number + 1
        else:
            new_number = 1001

        bill_number = f"FD-{new_number}"

        # 🔥 STEP 1: items lo
        item_names = request.POST.getlist("item_name[]")
        item_amounts = request.POST.getlist("item_amount[]")

        subtotal = 0
        doctor = None
        referred_by = None
        if referred_by_id:
            referred_by = UserProfile.objects.filter(
                id=referred_by_id,
                clinic=clinic,
                role__in=["owner", "doctor"]
            ).first()

        bill = Bill.objects.create(
            clinic=clinic,
            patient=patient,
              doctor=None,
            referred_by=referred_by,
            bill_number=bill_number,
            payment_mode=payment_mode,
        )

        # 🔥 STEP 5: save items
        for name, amount in zip(item_names, item_amounts):
            if name.strip() and amount:
                amount = float(amount)

                BillItem.objects.create(
                    bill=bill,
                    item_name=name,
                    amount=amount
                )
                subtotal += amount

        # 🔥 Discount calculation
        discount_amount = (subtotal * discount_percent) / 100
        total = subtotal - discount_amount

        bill.subtotal = subtotal
        bill.discount = discount_percent
        bill.discount_amount = discount_amount
        bill.total_amount = total

        bill.save()

        return redirect("view_bill", bill_id=bill.id)

    return render(request, "billing/create_bill.html", {
        "patients": patients,
        "clinic": clinic,
        "doctors": doctors,
        
    })


#Bill History
@login_required
def bill_history(request):
    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    
    if not has_permission(request.user, "manage_billing"):
        return render(request, "403.html", status=403)

    if not clinic.billing_enabled:
        return redirect("dashboard")

    bills = Bill.objects.filter(clinic=clinic).order_by("-created_at")

    # ========== FILTERS ==========
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    patient_id = request.GET.get('patient')
    doctor_id = request.GET.get('doctor')

    if start_date:
        bills = bills.filter(created_at__date__gte=start_date)
    if end_date:
        bills = bills.filter(created_at__date__lte=end_date)
    if patient_id:
        bills = bills.filter(patient_id=patient_id)
    if doctor_id:
        bills = bills.filter(doctor_id=doctor_id)
    
    # 🔥 PAGINATION START
    from django.core.paginator import Paginator

    per_page = request.GET.get("limit", 20)

    try:
        per_page = int(per_page)
    except:
        per_page = 20

    if per_page not in [20, 50]:
        per_page = 20

    paginator = Paginator(bills, per_page)
    page = request.GET.get("page")

    bills = paginator.get_page(page)
    # 🔥 PAGINATION END

    # ========== DROPDOWN DATA ==========
    patients = Patient.objects.filter(clinic=clinic).order_by('name')
    
    # Doctors = UserProfile with role owner/doctor
    doctors = UserProfile.objects.filter(
        clinic=clinic,
        role__in=["owner", "doctor"]
    ).order_by('name')

    # Selected names for active filter tags
    selected_patient_name = None
    selected_doctor_name = None
    
    if patient_id:
        selected_patient_name = patients.filter(id=patient_id).values_list('name', flat=True).first()
    if doctor_id:
        selected_doctor_name = doctors.filter(id=doctor_id).values_list('name', flat=True).first()

    return render(request, "billing/bill_history.html", {
        "bills": bills,
        "patients": patients,
        "doctors": doctors,
        "selected_patient_name": selected_patient_name,
        "selected_doctor_name": selected_doctor_name,
        "limit": per_page
    })

#Bill Detail
@login_required
def view_bill(request, bill_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not has_permission(request.user, "manage_billing"):
        return render(request, "403.html", status=403)

    bill = get_object_or_404(Bill, id=bill_id, clinic=clinic)

    items = bill.items.all()

    return render(request, "billing/view_bill.html", {
        "bill": bill,
        "items": items,
        "clinic": clinic,
        "hide_watermark": True
    })


  #Bill Print
@login_required
def print_bill(request, bill_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not has_permission(request.user, "manage_billing"):
        return render(request, "403.html", status=403)

    bill = get_object_or_404(
        Bill,
        id=bill_id,
        clinic=clinic
    )

    items = bill.items.all()

    return render(request, "billing/print_bill.html", {
        "bill": bill,
        "items": items,
        "clinic": clinic,
        "hide_watermark": True
    })


# Print Prescription
@login_required(login_url="login")
def print_prescription(request, id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    prescription = get_object_or_404(
        Prescription,
        id=id,
        clinic=clinic
    )

    # ✅ SAME LOGIC AS VIEW PRESCRIPTION
    med_lines = []

    if prescription.medicines:
        for med in prescription.medicines.split("\n"):
            if "||" in med:
                parts = med.split("||")
                med_lines.append({
                    "name": parts[0],
                    "dose": parts[1] if len(parts) > 1 else "",
                    "duration": parts[2] if len(parts) > 2 else "",
                    "remark": parts[3] if len(parts) > 3 else "",
                })
            else:
                med_lines.append({
                    "name": med,
                    "dose": "",
                    "duration": "",
                    "remark": "",
                })

    return render(request, "print_prescription.html", {
        "prescription": prescription,
        "clinic": clinic,
        "med_lines": med_lines,
        "hide_watermark": True   # 🔥 IMPORTANT
    })

@login_required
def enable_advanced_mode(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not profile.is_owner:
        return render(request, "403.html", status=403)

    if not clinic.is_advanced:
        clinic.is_advanced = True
        clinic.save()

    return redirect("dashboard")

from django.contrib.auth.models import User
from .models import UserProfile, UserPermission, Permission

@login_required
def add_staff(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not profile.is_owner:
        return render(request, "403.html", status=403)

    # 🔥 doctors list (for dropdown)
    doctors = UserProfile.objects.filter(
        clinic=clinic,
        role__in=["owner", "doctor"]
    )

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")
        name = request.POST.get("name")
        role = request.POST.get("role")
        email = request.POST.get("email")

        # ❗ Username validation
        if User.objects.filter(username=username).exists():
            return render(request, "staff/add_staff.html", {
                "error": "Username already exists",
                "doctors": doctors
            })

        # ✅ Create user
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email
        )

        # 🔥 Assistant doctor mapping
        assigned_doctor_id = request.POST.get("assigned_doctor")

        assigned_doctor = None
        if role == "assistant" and assigned_doctor_id:
            assigned_doctor = UserProfile.objects.filter(
                id=assigned_doctor_id,
                clinic=clinic,
                role__in=["owner", "doctor"]
            ).first()

        # ✅ Create profile
        profile_obj = UserProfile.objects.create(
            user=user,
            clinic=clinic,
            role=role,
            is_owner=False,
            name=name,
            assigned_doctor=assigned_doctor
        )

        # 🔥 DEFAULT PERMISSIONS AUTO ASSIGN

        if role == "doctor":
            default_perms = [
                "manage_patients",
                "manage_appointments",
                "create_prescription",
                "manage_billing"
            ]

            for perm_code in default_perms:
                perm = Permission.objects.filter(code=perm_code).first()
                if perm:
                    UserPermission.objects.create(
                        user_profile=profile_obj,
                        permission=perm
                    )

        elif role == "receptionist":
            default_perms = [
                "manage_patients",
                "manage_appointments",
                "manage_billing"
            ]

            for perm_code in default_perms:
                perm = Permission.objects.filter(code=perm_code).first()
                if perm:
                    UserPermission.objects.create(
                        user_profile=profile_obj,
                        permission=perm
                    )

        return redirect("staff_list")

    return render(request, "staff/add_staff.html", {
        "doctors": doctors
    })

@login_required
def staff_list(request):
    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not profile.is_owner:
        return render(request, "403.html", status=403)

    staff = UserProfile.objects.filter(clinic=clinic, is_owner=False, is_active=True)

    return render(request, "staff/staff_list.html", {
        "staff": staff
    })

from .models import UserPermission, Permission
@login_required
def edit_staff_permissions(request, staff_id):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    if not profile.is_owner:
        return render(request, "403.html", status=403)

    staff = get_object_or_404(UserProfile, id=staff_id, clinic=clinic)

    permissions = Permission.objects.all()

    # 🔥 CLEAN LABEL BANANE KE LIYE
    permissions_data = []
    for p in permissions:
        label = p.code.replace("_", " ").title()
        permissions_data.append({
            "code": p.code,
            "label": label
        })

    if request.method == "POST":

        selected_perms = request.POST.getlist("permissions")

        # Purane delete
        UserPermission.objects.filter(user_profile=staff).delete()

        # Naye add
        for perm_code in selected_perms:
            perm = Permission.objects.filter(code=perm_code).first()
            if perm:
                UserPermission.objects.create(
                    user_profile=staff,
                    permission=perm
                )

        return redirect("staff_list")

    # Existing permissions
    user_perms = UserPermission.objects.filter(user_profile=staff)
    user_perm_codes = [p.permission.code for p in user_perms]

    return render(request, "staff/edit_permissions.html", {
        "staff": staff,
        "permissions": permissions_data,   # 🔥 IMPORTANT CHANGE
        "user_perm_codes": user_perm_codes
    })

def staff_blocked(request):
    return render(request, 'staff_blocked.html')

@login_required
def export_all_bills(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not profile.is_owner:
        return render(request, "403.html", status=403)

    bills = Bill.objects.filter(clinic=clinic).select_related("patient", "doctor", "appointment")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bills"

    headers = [
        "Bill No",
        "Patient Name",
        "Phone",
        "Doctor",
        "Type",
        "Date",
        "Time",
        "Amount",
        "Payment Mode"
    ]

    sheet.append(headers)

    for bill in bills:

    # Doctor name
        doctor_name = "-"
        if bill.doctor:
            doctor_name = bill.doctor.name
        elif bill.referred_by:
            doctor_name = bill.referred_by.name + " (Referred)"

        # Type
        bill_type = "Consultation" if bill.doctor else "Service"

        # Appointment time
        appointment_time = "-"
        if bill.appointment and bill.appointment.appointment_time:
            appointment_time = bill.appointment.appointment_time.strftime("%H:%M")

        sheet.append([
            bill.bill_number,
            bill.patient.name,
            bill.patient.phone,
            doctor_name,
            bill_type,
            bill.created_at.strftime("%d-%m-%Y"),
            appointment_time,
            bill.total_amount,
            bill.payment_mode
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="all_bills.xlsx"'

    workbook.save(response)

    return response

@login_required
def export_month_bills(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic

    if not profile.is_owner:
        return render(request, "403.html", status=403)

    month = request.GET.get("month")
    year = request.GET.get("year")

    if not month or not year:
        today = date.today()
        month = today.month
        year = today.year

    bills = Bill.objects.filter(
        clinic=clinic,
        created_at__month=int(month),
        created_at__year=int(year)
    ).select_related("patient", "doctor", "appointment")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bills"

    headers = [
        "Bill No",
        "Patient Name",
        "Phone",
        "Doctor",
        "Type",
        "Date",
        "Time",
        "Amount",
        "Payment Mode"
    ]

    sheet.append(headers)

    for bill in bills:

        # Doctor name
        doctor_name = "-"
        if bill.doctor:
            doctor_name = bill.doctor.name
        elif bill.referred_by:
            doctor_name = bill.referred_by.name + " (Referred)"

        # Type
        bill_type = "Consultation" if bill.doctor else "Service"

        # Appointment time
        appointment_time = "-"
        if bill.appointment and bill.appointment.appointment_time:
            appointment_time = bill.appointment.appointment_time.strftime("%H:%M")

        sheet.append([
            bill.bill_number,
            bill.patient.name,
            bill.patient.phone,
            doctor_name,
            bill_type,
            bill.created_at.strftime("%d-%m-%Y"),
            appointment_time,
            bill.total_amount,
            bill.payment_mode
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"bills_{month}_{year}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response


from django.utils import timezone

def get_queue_data(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    today = timezone.localtime().date()

    # 🔥 SAME FILTER AS DASHBOARD
    if profile.role in ["doctor", "owner"]:
        appointments = Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            doctor=profile
        )

    elif profile.role == "assistant":
        if profile.assigned_doctor:
            appointments = Appointment.objects.filter(
                clinic=clinic,
                appointment_date=today,
                doctor=profile.assigned_doctor
            )
        else:
            appointments = Appointment.objects.none()

    else:  # receptionist
        appointments = Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today
        )

    appointments = appointments.order_by("doctor", "token_number")

    data = []

    for a in appointments:
        data.append({
            "id": a.id,
            "patient_id": a.patient.id,
            "status": a.status,
            "queue_status": a.queue_status,
            "token_number": a.token_number,
        })

    # 🔥 SAME NEXT LOGIC (copy from dashboard)
    busy_doctors = set(
        Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            queue_status="in_consultation"
        ).values_list("doctor_id", flat=True)
    )

    all_waiting = Appointment.objects.filter(
        clinic=clinic,
        appointment_date=today,
        queue_status="waiting"
    ).order_by("doctor", "token_number")

    next_tokens = []
    seen_doctors = set()

    for appt in all_waiting:
        if appt.doctor_id not in seen_doctors and appt.doctor_id not in busy_doctors:
            next_tokens.append(appt.id)
            seen_doctors.add(appt.doctor_id)

    return JsonResponse({
        "appointments": data,
        "next_tokens": next_tokens
    })


@login_required
@require_http_methods(["GET"])
def api_queue(request):

    profile = get_object_or_404(UserProfile, user=request.user)
    clinic = profile.clinic
    today = datetime.now().date()

    # 🔥 SAME LOGIC AS DASHBOARD
    if profile.role in ["doctor", "owner"]:
        appointments = Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today,
            doctor=profile
        )

    elif profile.role == "assistant":
        if profile.assigned_doctor:
            appointments = Appointment.objects.filter(
                clinic=clinic,
                appointment_date=today,
                doctor=profile.assigned_doctor
            )
        else:
            appointments = Appointment.objects.none()

    else:
        appointments = Appointment.objects.filter(
            clinic=clinic,
            appointment_date=today
        )

    total_appointments = appointments.count()
    pending_count = appointments.filter(status="pending").count()
    in_consultation_count = appointments.filter(queue_status="in_consultation").count()
    completed_count = appointments.filter(status="completed").count()

    return JsonResponse({
        "total_appointments": total_appointments,
        "pending_count": pending_count,
        "in_consultation_count": in_consultation_count,
        "completed_count": completed_count,
    })


from django.views.decorators.http import require_POST

@require_POST
@login_required
def deactivate_staff(request, staff_id):
    profile = request.user.userprofile

    # 🔴 Only owner allowed
    if not profile.is_owner:
        return redirect("dashboard")

    # 🔴 Same clinic ka staff hi mile
    staff = get_object_or_404(UserProfile, id=staff_id, clinic=profile.clinic)

    # ❌ Owner ko remove nahi karna
    if staff.is_owner:
        return redirect("staff_list")

    # ✅ MAIN LOGIC
    staff.is_active = False
    staff.save()

    return redirect("staff_list")

def revenue_report(request):

    clinic = request.user.userprofile.clinic

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    doctor_id = request.GET.get("doctor")

    today = timezone.localdate()

    # 🔥 DEFAULT = TODAY
    bills = Bill.objects.filter(
        clinic=clinic,
        created_at__date=today
    )

    # 🔥 FILTER override
    if start_date or end_date or doctor_id:

        bills = Bill.objects.filter(clinic=clinic)

        if start_date:
            bills = bills.filter(created_at__date__gte=start_date)

        if end_date:
            bills = bills.filter(created_at__date__lte=end_date)

        if doctor_id:
            bills = bills.filter(doctor_id=doctor_id)

    total = bills.aggregate(total=Sum("total_amount"))["total"] or 0
    

    doctor_revenue = BillItem.objects.filter(
        bill__in=bills,
        bill__clinic=clinic,
        bill__doctor__isnull=False,   # ❗ NA hata diya
        item_name__icontains="consultation"
    ).values(
        "bill__doctor__id",
        "bill__doctor__name"
    ).annotate(
        total=Sum("amount")
    ).order_by("-total")
    doctor_revenue = list(doctor_revenue)

    total_doctor_revenue = sum(d["total"] for d in doctor_revenue)
    clinic_revenue = total - total_doctor_revenue

    for d in doctor_revenue:
        d["percent"] = (
            (d["total"] / total_doctor_revenue) * 100
            if total_doctor_revenue > 0 else 0
        )


    doctors = UserProfile.objects.filter(
        clinic=clinic,
        role__in=["owner", "doctor"]
    )
    
    bill_count = bills.count()
    avg_bill = total / bill_count if bill_count > 0 else 0

    return render(request, "revenue_report.html", {
        "doctor_revenue": doctor_revenue,
        "total": total,
        "doctors": doctors,
        "bill_count": bill_count,     
        "avg_bill": avg_bill,          
        "today": today,
        "total_doctor_revenue": total_doctor_revenue,      
        "clinic_revenue": clinic_revenue,          
    })

from django.db.models import Q
from .models import Medicine

@login_required
def search_medicine(request):
    query = request.GET.get('q', '').strip()

    if not query:
        return JsonResponse([], safe=False)

    clinic = request.user.userprofile.clinic

    medicines = Medicine.objects.filter(
        Q(clinic=clinic) | Q(clinic__isnull=True),
        name__istartswith=query
    ).order_by('-usage_count')[:10]

    return JsonResponse([m.name for m in medicines], safe=False)

def save_medicine(name, clinic):
    name = name.strip()

    if not name:
        return

    med, created = Medicine.objects.get_or_create(
        name__iexact=name,
        clinic=clinic,
        defaults={"name": name}
    )

    med.usage_count += 1
    med.save(update_fields=["usage_count"])
